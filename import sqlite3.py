import sqlite3
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import openpyxl
from openpyxl import Workbook
import os

# Connect to SQLite database (or create one)
conn = sqlite3.connect('quiz_app.db')
cursor = conn.cursor()

# Create the questions table if it doesn't exist
cursor.execute(''' 
CREATE TABLE IF NOT EXISTS questions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    question TEXT NOT NULL,
    option1 TEXT NOT NULL,
    option2 TEXT NOT NULL,
    option3 TEXT NOT NULL,
    option4 TEXT NOT NULL,
    correct_option INTEGER NOT NULL,
    category TEXT NOT NULL
)
''')

# Sample questions
sample_questions = [
    ("What is the capital of France?", "Berlin", "Madrid", "Paris", "Rome", 3, "Geography"),
    ("Who developed Python?", "Ryan Gosling", "Guido van Rossum", "Brendan Eich", "Bjarne Stroustrup", 2, "Technology"),
    ("What is the largest planet in our solar system?", "Earth", "Mars", "Jupiter", "Saturn", 3, "Science"),
]

# Insert sample questions if table is empty
cursor.execute("SELECT COUNT(*) FROM questions")
if cursor.fetchone()[0] == 0:
    cursor.executemany(''' 
        INSERT INTO questions (question, option1, option2, option3, option4, correct_option, category)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', sample_questions)
    conn.commit()

# Tkinter Quiz App class
class QuizApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Quiz App")
        self.root.geometry("800x600")

        # Initialize quiz variables
        self.score = 0
        self.question_index = 0
        self.questions = []
        self.time_limit = 10
        self.time_remaining = self.time_limit
        self.username = ""

        # Fetch categories
        cursor.execute('SELECT DISTINCT category FROM questions')
        self.categories = [row[0] for row in cursor.fetchall()]

        # Ask for the username before the quiz
        self.ask_for_username()

    def ask_for_username(self):
        """Ask for the user's name before starting the quiz"""
        self.username_label = tk.Label(self.root, text="Enter your name:", font=("Arial", 16, "italic"), bg="lightblue")
        self.username_label.pack(pady=10)

        self.username_entry = tk.Entry(self.root, font=("Arial", 16))
        self.username_entry.pack(pady=10)

        self.start_button = tk.Button(self.root, text="Start Quiz", command=self.start_quiz, font=("Arial", 16, "italic"))
        self.start_button.pack(pady=20)

    def start_quiz(self):
        """Start the quiz after category selection"""
        self.username = self.username_entry.get().strip()
        if not self.username:
            messagebox.showwarning("Input Error", "Please enter a valid name.")
            return

        # Hide the username prompt and show the category selection screen
        self.username_label.pack_forget()
        self.username_entry.pack_forget()
        self.start_button.pack_forget()

        # Category selection UI
        self.category_label = tk.Label(self.root, text="Choose a category:", font=("Arial", 16, "bold"), bg="green")
        self.category_label.pack(pady=10)

        self.category_var = tk.StringVar(value=self.categories[0])
        self.category_menu = tk.OptionMenu(self.root, self.category_var, *self.categories)
        self.category_menu.pack(pady=10)

        self.time_limit_label = tk.Label(self.root, text="Set time limit per question (seconds):", font=("Arial", 16, "bold"), bg="cyan")
        self.time_limit_label.pack(pady=10)

        self.time_limit_entry = tk.Entry(self.root)
        self.time_limit_entry.insert(0, "10")
        self.time_limit_entry.pack(pady=10)

        self.start_button_quiz = tk.Button(self.root, text="Start Quiz", command=self.begin_quiz, font=("Arial", 16, "bold"), bg="blue")
        self.start_button_quiz.pack(pady=10)

    def begin_quiz(self):
        """Start the quiz after category selection"""
        selected_category = self.category_var.get()
        cursor.execute('SELECT * FROM questions WHERE category = ?', (selected_category,))
        self.questions = cursor.fetchall()

        # Customize time limit
        try:
            self.time_limit = max(5, min(60, int(self.time_limit_entry.get())))
        except ValueError:
            self.time_limit = 10

        self.question_index = 0
        self.score = 0
        self.time_remaining = self.time_limit

        self.category_label.pack_forget()
        self.category_menu.pack_forget()
        self.time_limit_label.pack_forget()
        self.time_limit_entry.pack_forget()
        self.start_button_quiz.pack_forget()

        self.show_quiz_widgets()
        self.load_question()
        self.update_timer()

    def show_quiz_widgets(self):
        """Show quiz widgets when the quiz starts"""
        self.question_label = tk.Label(self.root, text="", wraplength=300, font=("Arial", 16, "bold"))
        self.question_label.pack(pady=20)

        self.var = tk.IntVar()

        self.radio1 = tk.Radiobutton(self.root, text="", variable=self.var, value=1, font=("Arial", 16))
        self.radio1.pack(anchor="w")
        self.radio2 = tk.Radiobutton(self.root, text="", variable=self.var, value=2, font=("Arial", 16))
        self.radio2.pack(anchor="w")
        self.radio3 = tk.Radiobutton(self.root, text="", variable=self.var, value=3, font=("Arial", 16))
        self.radio3.pack(anchor="w")
        self.radio4 = tk.Radiobutton(self.root, text="", variable=self.var, value=4, font=("Arial", 16))
        self.radio4.pack(anchor="w")

        self.submit_button = tk.Button(self.root, text="Submit", command=self.check_answer, font=("Arial", 16))
        self.submit_button.pack(pady=20)

        self.timer_label = tk.Label(self.root, text="", font=("Arial", 16, "bold"))
        self.timer_label.pack(pady=10)

    def load_question(self):
        """Load a question from the database"""
        if self.question_index < len(self.questions):
            question = self.questions[self.question_index]
            self.question_label.config(text=question[1])
            self.radio1.config(text=question[2])
            self.radio2.config(text=question[3])
            self.radio3.config(text=question[4])
            self.radio4.config(text=question[5])
            self.var.set(0)
            self.time_remaining = self.time_limit
        else:
            self.show_score()

    def check_answer(self):
        """Check the selected answer and update score"""
        if self.var.get() == self.questions[self.question_index][6]:
            self.score += 1
        self.question_index += 1
        self.load_question()

    def update_timer(self):
        """Update the timer for each question"""
        if self.time_remaining > 0:
            self.timer_label.config(text=f"Time remaining: {self.time_remaining} seconds")
            self.time_remaining -= 1
            self.root.after(1000, self.update_timer)
        else:
            self.check_answer()

    def show_score(self):
        """Show the final score and save to Excel"""
        messagebox.showinfo("Quiz Over", f"Your score: {self.score}/{len(self.questions)}")
        self.save_to_excel()

    def save_to_excel(self):
        """Save the score and category to an Excel file"""
        file_exists = os.path.isfile('quiz_scores.xlsx')
        if file_exists:
            wb = openpyxl.load_workbook('quiz_scores.xlsx')
        else:
            wb = Workbook()

        ws = wb.active
        if ws.max_row == 1:
            ws.append(["Username", "Score", "Total Questions", "Category"])
        ws.append([self.username, self.score, len(self.questions), self.category_var.get()])
        wb.save('quiz_scores.xlsx')

# Create the main Tkinter window
root = tk.Tk()
root.title("Quiz App")
root.geometry("800x600")

# Load and set background image
background_image = Image.open("C:\Users\killi\Downloads\a6a34faa-40ba-406a-af33-df4933cc6485.jpg")  # Replace with your image path
background_image = background_image.resize((800, 600), Image.ANTIALIAS)
bg_photo = ImageTk.PhotoImage(background_image)
bg_label = tk.Label(root, image=bg_photo)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

# Start QuizApp
quiz_app = QuizApp(root)
root.mainloop()
